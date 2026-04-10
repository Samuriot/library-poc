from typing import List
from collections import defaultdict
from datetime import datetime
import csv
import os
import sys
import re
from pathlib import Path
import openpyxl

def import_dataset_xlsx(filename: str) -> dict[str, List]:
    locations = defaultdict(list)
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Read headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Process data rows (starting from row 2)
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(value for value in row if value):
            continue
        
        # Create a dictionary for this row
        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        
        date = row_dict.get('Date', '')
        time = row_dict.get('Time', '')
        
        # Convert date to YYYY-MM-DD format
        date_formatted = date
        if isinstance(date, datetime):
            date_formatted = date.strftime('%Y-%m-%d')
        elif isinstance(date, str):
            for date_fmt in ['%m/%d/%Y', '%m/%d/%y']:
                try:
                    date_obj = datetime.strptime(date, date_fmt)
                    date_formatted = date_obj.strftime('%Y-%m-%d')
                    break
                except:
                    continue
        
        # Convert time from 12-hour to 24-hour format
        if isinstance(time, datetime):
            time_24hr = time.strftime('%H:%M')
        elif isinstance(time, str):
            try:
                time_obj = datetime.strptime(time, '%I:%M %p')
                time_24hr = time_obj.strftime('%H:%M')
            except:
                time_24hr = time
        else:
            time_24hr = str(time) if time else ''
        
        datetime_str = f"{date_formatted} {time_24hr}"
        
        # Process each location column
        for item in row_dict:
            if item in ['Time', 'Date']:
                continue
            value = row_dict[item]
            count_value = str(value).strip() if value is not None else "-1"
            entry_data = {
                'datetime': datetime_str,
                'count': int(float(count_value)) if count_value and count_value != '' else 0
            }
            locations[item].append(entry_data)
    
    locations_dict: dict[str, List] = {location: entries for location, entries in locations.items()}
    return locations_dict

def import_dataset_csv(filename: str) -> dict[str, List]:
    locations = defaultdict(list)
    with open(filename, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:            
            if not any((value or "").strip() for value in row.values()):
                continue
            row.pop("", None)
            date = row.get('Date', '')
            time = row.get('Time', '')
            
            # Convert date to YYYY-MM-DD format
            date_formatted = date
            for date_fmt in ['%m/%d/%Y', '%m/%d/%y']:
                try:
                    date_obj = datetime.strptime(date, date_fmt)
                    date_formatted = date_obj.strftime('%Y-%m-%d')
                    break
                except:
                    continue
            
            # Convert time from 12-hour to 24-hour format
            try:
                time_obj = datetime.strptime(time, '%I:%M %p')
                time_24hr = time_obj.strftime('%H:%M')
            except:
                time_24hr = time
            
            datetime_str = f"{date_formatted} {time_24hr}"
            
            for item in row:
                if item in ['Time', 'Date']:
                    continue
                count_value = row[item].strip() if row[item] else "-1"
                entry_data = {
                    'datetime': datetime_str,
                    'count': int(count_value) if count_value else 0
                }
                locations[item].append(entry_data)
    
    locations_dict: dict[str, List] = {location: entries for location, entries in locations.items()}
    return locations_dict


def validate_data(locations_dict):    
    # Regex pattern for YYYY-MM-DD format
    date_pattern = re.compile(r'^\d{4}-\d{2}-\d{2}')
    
    for location, entries in locations_dict.items():
        for entry in entries:
            # Validate date format (YYYY-MM-DD)
            datetime_str = entry.get('datetime', '')
            if not date_pattern.match(datetime_str):
                # Try to fix common date formats
                try:
                    # Extract date part before the time
                    parts = datetime_str.split(' ')
                    if len(parts) >= 2:
                        date_part = parts[0]
                        time_part = ' '.join(parts[1:])
                        
                        # Try various date formats
                        for fmt in ['%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y', '%d/%m/%y', '%Y/%m/%d']:
                            try:
                                date_obj = datetime.strptime(date_part, fmt)
                                entry['datetime'] = f"{date_obj.strftime('%Y-%m-%d')} {time_part}"
                                break
                            except:
                                continue
                except:
                    pass
            
            # Validate count is not negative
            if entry.get('count', 0) < 0:
                entry['count'] = 0
    
    return locations_dict


def export_to_csv(locations_dict):
    # Create data directory if it doesn't exist
    os.makedirs('data', exist_ok=True)
    
    for location, entries in locations_dict.items():
        # Create a safe filename from location name
        filename = f"data/{location.replace(' ', '_').replace('/', '_')}.csv"
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # Write header
            writer.writerow(['date', 'gate_start'])
            
            # Write data rows
            for entry in entries:
                writer.writerow([entry['datetime'], entry['count']])
        
        print(f"Exported {filename}")


def main():
    if len(sys.argv) < 2:
        print("Exiting program, provide the csv/xlsx file name to parse - E.G. python3 parse.py <FILE_NAME>.csv")
        return
    file_path = Path(sys.argv[1])
    if file_path.suffix not in [".csv", ".xlsx"]:
        print("Exiting program, file provided is not a .csv or .xlsx - E.G. python3 parse.py <FILE_NAME>.csv")
        return
    
    # Choose the appropriate import function based on file extension
    if file_path.suffix == ".csv":
        formatted_dataset = import_dataset_csv(file_path.name)
    else:  # .xlsx
        formatted_dataset = import_dataset_xlsx(file_path.name)
    
    validated_dataset = validate_data(formatted_dataset)
    export_to_csv(validated_dataset)

if __name__ == '__main__':
    main()
